# -*- coding: utf-8 -*-
"""
╔══════════════════════════════════════════════════════════════════════════════╗
║   Multi-Modal Autonomous Data Entry Assistant                               ║
║   LangGraph  ·  Human-in-the-Loop  ·  Groq llama-3.3-70b-versatile         ║
╠══════════════════════════════════════════════════════════════════════════════╣
║                                                                              ║
║  GRAPH ARCHITECTURE                                                          ║
║  ──────────────────                                                          ║
║                                                                              ║
║  [START]                                                                     ║
║     │                                                                        ║
║     ▼                                                                        ║
║  ┌──────────────┐    conditional edge                                        ║
║  │ detect_type  │──────────────────────────────────────────┐                 ║
║  └──────────────┘                                          │                 ║
║        │ (4 branches)                                      │                 ║
║   ┌────┴─────┬──────────┬──────────┐                       │                 ║
║   ▼          ▼          ▼          ▼                       │                 ║
║ [img]     [pdf]      [audio]    [text]                     │                 ║
║ loader    loader     loader     loader                     │                 ║
║   └────┬─────┴──────────┴──────────┘                       │                 ║
║        │  (converge)                                        │                 ║
║        ▼                                                    │                 ║
║  ┌──────────────┐                                           │                 ║
║  │ extract_data │  llama-3.3-70b extracts schema + records  │                 ║
║  └──────┬───────┘                                           │                 ║
║         │  conditional edge (error → report_error)          │                 ║
║         ▼                                                    │                 ║
║  ┌──────────────┐  ◄── CHECKPOINT 1 ────────────────────────┘                 ║
║  │ human_review │  Human sees preview, approves or edits                      ║
║  └──────┬───────┘                                                             ║
║         │  conditional edge                                                   ║
║    ┌────┴──────┐                                                              ║
║    ▼           ▼                                                              ║
║ [approved]  [rejected]                                                       ║
║    │              │                                                           ║
║    ▼              ▼                                                           ║
║ save_outputs  re_extract  ◄── CHECKPOINT 2 (retry loop)                      ║
║    │              │                                                           ║
║    ▼              └─────────────────► human_review (loop back)               ║
║ ┌──────────────┐                                                              ║
║ │    report    │  Final summary                                               ║
║ └──────┬───────┘                                                              ║
║        ▼                                                                      ║
║      [END]                                                                    ║
║                                                                               ║
╠═══════════════════════════════════════════════════════════════════════════════╣
║  HUMAN-IN-THE-LOOP CHECKPOINTS                                                ║
║  • interrupt_before=[\"human_review\"] pauses the graph after extraction       ║
║  • The user sees a data preview and can: approve / edit / reject              ║
║  • Graph resumes from the checkpoint via graph.invoke(Command(resume=...))   ║
║  • MemorySaver persists full state between pause and resume                  ║
╚═══════════════════════════════════════════════════════════════════════════════╝

Usage
─────
  # 1. Create a .env file in the same folder as this script:
  #       GROQ_API_KEY=gsk_your-key-here
  #
  # 2. Run — the key is loaded automatically from .env

  # Interactive HITL mode (recommended)
  python multimodal_data_entry_assistant.py

  # File inputs
  python multimodal_data_entry_assistant.py -i invoice.pdf
  python multimodal_data_entry_assistant.py -i receipt.jpg
  python multimodal_data_entry_assistant.py -i contacts.txt
  python multimodal_data_entry_assistant.py -i meeting_transcript.txt

  # Skip human review (fully autonomous)
  python multimodal_data_entry_assistant.py -i data.pdf --auto-approve

  # Multiple files
  python multimodal_data_entry_assistant.py -i a.pdf b.jpg c.txt
"""

from __future__ import annotations

import argparse
import base64
import csv
import json
import logging
import mimetypes
import os
import sys
import uuid
from datetime import datetime
from functools import lru_cache
from pathlib import Path
from typing import Any, Literal

# ── dotenv ────────────────────────────────────────────────────────────────────
try:
    from dotenv import load_dotenv
    _env_path = Path(__file__).parent / ".env"
    if _env_path.exists():
        load_dotenv(dotenv_path=_env_path, override=False)
        print(f"  Loaded .env from: {_env_path}")
    else:
        load_dotenv(override=False)
except ImportError:
    print(
        "  WARNING: python-dotenv not installed — .env will NOT be loaded.\n"
        "  Run:  pip install -r requirements.txt\n"
    )

# ── LangGraph ─────────────────────────────────────────────────────────────────
try:
    from langgraph.graph import StateGraph, END
    from langgraph.checkpoint.memory import MemorySaver
    from langgraph.types import Command, interrupt
    from typing_extensions import TypedDict, Annotated
    import operator
except ImportError:
    sys.exit(
        "\n[ERROR] LangGraph not found.\n"
        "Run:  pip install -r requirements.txt\n"
    )

# ── Groq ──────────────────────────────────────────────────────────────────────
try:
    from groq import Groq
except ImportError:
    sys.exit("\n[ERROR] groq not found.\nRun: pip install -r requirements.txt\n")

# ── Optional deps ─────────────────────────────────────────────────────────────
try:
    import fitz
    PDF_OK = True
except ImportError:
    PDF_OK = False

try:
    from PIL import Image
    import io as _io
    PIL_OK = True
except ImportError:
    PIL_OK = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

# ── Logging ───────────────────────────────────────────────────────────────────
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  [%(levelname)-8s]  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)

# ── Constants ─────────────────────────────────────────────────────────────────
GROQ_MODEL        = "llama-3.3-70b-versatile"
GROQ_VISION_MODEL = "meta-llama/llama-4-scout-17b-16e-instruct"  # Groq vision model
GROQ_AUDIO_MODEL  = "whisper-large-v3"                            # Groq Whisper endpoint

IMAGE_EXTS = {".jpg", ".jpeg", ".png", ".gif", ".webp", ".bmp", ".tiff"}
AUDIO_EXTS = {".wav", ".mp3", ".m4a", ".flac", ".ogg", ".mp4", ".webm"}
PDF_EXTS   = {".pdf"}
TEXT_EXTS  = {".txt", ".md", ".csv", ".json", ".xml", ".html", ".log", ".tsv"}

EXTRACTION_PROMPT = """You are an expert autonomous data extraction agent.

Analyse the input and extract ALL structured data from it.

STRICT RULES:
1. Return ONLY valid JSON — no markdown fences, no prose.
2. Two top-level keys only:
   "schema"  — list of snake_case column name strings
   "records" — list of objects; every object must contain ALL schema keys
3. Missing values → null (never omit a key).
4. Dates → YYYY-MM-DD.  Monetary → float.  Booleans → true/false.
5. Be exhaustive — extract every row you can find.

Example:
{"schema":["name","email","amount","date"],
 "records":[{"name":"Alice","email":"a@x.com","amount":99.0,"date":"2024-03-15"}]}
"""


# ══════════════════════════════════════════════════════════════════════════════
#  GROQ CLIENT (cached per API key)
# ══════════════════════════════════════════════════════════════════════════════

@lru_cache(maxsize=4)
def _get_groq_client(api_key: str) -> Groq:
    return Groq(api_key=api_key)


def _groq(state: dict | None = None) -> Groq:
    """Return a cached Groq client, pulling key from state or env."""
    key = (state or {}).get("api_key") or os.environ.get("GROQ_API_KEY", "")
    if not key:
        raise EnvironmentError("GROQ_API_KEY is not set.")
    return _get_groq_client(key)


# ══════════════════════════════════════════════════════════════════════════════
#  STATE
# ══════════════════════════════════════════════════════════════════════════════

class AgentState(TypedDict):
    # Pipeline config (set at start, never mutated)
    source:       str
    output_dir:   str
    formats:      list[str]
    api_key:      str | None
    auto_approve: bool

    # Set by detect_type
    input_type:   str                        # "image" | "pdf" | "audio" | "text"

    # Set by loaders — unified message list sent to the LLM
    groq_messages: list[dict]

    # Set by extract_data / re_extract
    schema:             list[str]
    records:            list[dict]
    extraction_error:   str | None

    # Set by human_review (HITL checkpoint)
    review_decision:    Literal["approved", "rejected", "edited"] | None
    human_feedback:     str | None
    edited_records:     list[dict] | None

    # Append-only audit log
    messages: Annotated[list[str], operator.add]

    # Retry counter (reducer adds)
    retries: Annotated[int, operator.add]

    # Set by save_outputs
    output_files: dict[str, str]

    # Set by report
    summary: dict


# ══════════════════════════════════════════════════════════════════════════════
#  NODES
# ══════════════════════════════════════════════════════════════════════════════

# ── 1. detect_type ────────────────────────────────────────────────────────────
def detect_type(state: AgentState) -> dict:
    p   = Path(state["source"])
    ext = p.suffix.lower() if p.exists() else ""

    if not p.exists():
        t = "text"
    elif ext in IMAGE_EXTS:
        t = "image"
    elif ext in PDF_EXTS:
        t = "pdf"
    elif ext in AUDIO_EXTS:
        t = "audio"
    else:
        t = "text"

    log.info("detect_type → %s", t)
    return {
        "input_type": t,
        "messages":   [f"[detect_type] source='{state['source']}' type={t}"],
    }


# ── 2a. load_image ────────────────────────────────────────────────────────────
def load_image(state: AgentState) -> dict:
    """
    Encode the image as base64 and build a vision message for
    Groq's multimodal endpoint (llama-4-scout).
    """
    path = Path(state["source"])
    log.info("load_image: %s", path.name)

    mime, _ = mimetypes.guess_type(str(path))
    supported = {"image/jpeg", "image/png", "image/gif", "image/webp"}

    if mime not in supported and PIL_OK:
        img = Image.open(path).convert("RGB")
        buf = _io.BytesIO()
        img.save(buf, format="PNG")
        data = base64.standard_b64encode(buf.getvalue()).decode()
        mime = "image/png"
    else:
        data = base64.standard_b64encode(path.read_bytes()).decode()
        mime = mime or "image/jpeg"

    msgs = [{
        "role": "user",
        "content": [
            {
                "type": "image_url",
                "image_url": {"url": f"data:{mime};base64,{data}"},
            },
            {
                "type": "text",
                "text": f"Image: {path.name}\nExtract ALL structured data. Return JSON only.",
            },
        ],
    }]
    return {
        "groq_messages": msgs,
        "messages": [f"[load_image] encoded {path.name}"],
    }


# ── 2b. load_pdf ──────────────────────────────────────────────────────────────
def load_pdf(state: AgentState) -> dict:
    if not PDF_OK:
        raise RuntimeError("PyMuPDF not installed. Run: pip install pymupdf")
    path = Path(state["source"])
    log.info("load_pdf: %s", path.name)

    doc   = fitz.open(str(path))
    pages = [
        f"--- Page {i} ---\n{pg.get_text().strip()}"
        for i, pg in enumerate(doc, 1)
        if pg.get_text().strip()
    ]
    doc.close()

    if not pages:
        raise ValueError(f"No readable text in: {path.name}")

    msgs = [{
        "role": "user",
        "content": (
            f"PDF: {path.name}\n\n"
            + "\n\n".join(pages)
            + "\n\nExtract ALL structured data. Return JSON only."
        ),
    }]
    return {
        "groq_messages": msgs,
        "messages": [f"[load_pdf] {len(pages)} pages"],
    }


# ── 2c. load_audio ────────────────────────────────────────────────────────────
def load_audio(state: AgentState) -> dict:
    """
    Transcribe audio using Groq's Whisper endpoint, then build a text
    message for the extraction step.
    """
    path = Path(state["source"])
    log.info("load_audio: transcribing %s with Groq Whisper…", path.name)

    client = _groq(state)
    with open(path, "rb") as f:
        transcription = client.audio.transcriptions.create(
            model=GROQ_AUDIO_MODEL,
            file=f,
            response_format="text",
        )

    transcript_text = transcription if isinstance(transcription, str) else str(transcription)

    msgs = [{
        "role": "user",
        "content": (
            f"Audio: {path.name}\nWhisper transcript:\n\n{transcript_text}"
            "\n\nExtract ALL structured data. Return JSON only."
        ),
    }]
    return {
        "groq_messages": msgs,
        "messages": [f"[load_audio] transcript len={len(transcript_text)}"],
    }


# ── 2d. load_text ─────────────────────────────────────────────────────────────
def load_text(state: AgentState) -> dict:
    p = Path(state["source"])
    if p.exists():
        content, label = p.read_text(encoding="utf-8", errors="replace"), f"file '{p.name}'"
    else:
        content, label = state["source"], "typed/clipboard text"

    log.info("load_text: %s (%d chars)", label, len(content))
    msgs = [{
        "role": "user",
        "content": (
            f"Source: {label}\n\n{content}"
            "\n\nExtract ALL structured data. Return JSON only."
        ),
    }]
    return {
        "groq_messages": msgs,
        "messages": [f"[load_text] {label}"],
    }


# ── 3. extract_data ───────────────────────────────────────────────────────────
def extract_data(state: AgentState) -> dict:
    retry = state.get("retries", 0)
    log.info("extract_data: calling %s (retry=%d)…", GROQ_MODEL, retry)

    # Inject human feedback if this is a re-extraction
    extra = ""
    if state.get("human_feedback"):
        extra = (
            f"\n\nIMPORTANT CORRECTION FROM HUMAN REVIEWER:\n{state['human_feedback']}"
            "\nPlease fix the extraction accordingly."
        )

    messages = state["groq_messages"].copy()
    if extra:
        last = messages[-1]
        if isinstance(last["content"], str):
            messages[-1] = {**last, "content": last["content"] + extra}
        else:
            # vision message: append extra as a new text block
            messages[-1] = {
                **last,
                "content": last["content"] + [{"type": "text", "text": extra}],
            }

    # Choose model: vision model if content contains image_url, else text model
    has_image = any(
        isinstance(m.get("content"), list)
        and any(b.get("type") == "image_url" for b in m["content"])
        for m in messages
    )
    model = GROQ_VISION_MODEL if has_image else GROQ_MODEL

    client = _groq(state)
    try:
        resp = client.chat.completions.create(
            model=model,
            max_tokens=4096,
            response_format={"type": "json_object"},
            messages=[{"role": "system", "content": EXTRACTION_PROMPT}, *messages],
        )
        raw     = resp.choices[0].message.content.strip()
        parsed  = json.loads(raw)
        schema  = parsed.get("schema", [])
        records = parsed.get("records", [])
        log.info("Extracted %d cols × %d rows (model=%s)", len(schema), len(records), model)
        return {
            "schema":           schema,
            "records":          records,
            "extraction_error": None,
            "human_feedback":   None,   # clear after use
            "messages": [f"[extract_data] model={model} {len(schema)} cols, {len(records)} rows"],
        }
    except Exception as exc:
        log.error("Extraction failed: %s", exc)
        return {
            "schema":           [],
            "records":          [],
            "extraction_error": str(exc),
            "messages":         [f"[extract_data] ERROR: {exc}"],
        }


# ── 4. human_review  (HITL CHECKPOINT) ───────────────────────────────────────
def human_review(state: AgentState) -> Command:
    """
    HUMAN-IN-THE-LOOP checkpoint.

    Uses LangGraph's interrupt() to pause execution and surface a preview
    of the extracted data to the human operator.

    Responses accepted:
      • {"decision": "approved"}
      • {"decision": "rejected", "feedback": "fix the dates…"}
      • {"decision": "edited",   "records": [...corrected rows...]}
    """
    if state.get("auto_approve"):
        log.info("human_review: auto-approved")
        return Command(update={
            "review_decision": "approved",
            "messages": ["[human_review] auto-approved"],
        })

    schema  = state.get("schema", [])
    records = state.get("records", [])

    # ── Terminal preview ──────────────────────────────────────────────────────
    preview_lines = [
        "",
        "┌─────────────────────────────────────────────────────────────┐",
        "│              HUMAN REVIEW  —  CHECKPOINT                    │",
        "├─────────────────────────────────────────────────────────────┤",
        f"│  Source     : {state['source'][:55]:<55}│",
        f"│  Input type : {state.get('input_type', '?'):<55}│",
        f"│  Model      : {GROQ_MODEL:<55}│",
        f"│  Columns    : {len(schema):<55}│",
        f"│  Records    : {len(records):<55}│",
        "├─────────────────────────────────────────────────────────────┤",
        "│  SCHEMA                                                      │",
    ]
    for col in schema:
        preview_lines.append(f"│    • {col:<57}│")
    preview_lines.append("├─────────────────────────────────────────────────────────────┤")
    preview_lines.append("│  DATA PREVIEW (first 5 rows)                                │")

    for i, rec in enumerate(records[:5], 1):
        preview_lines.append(f"│  Row {i}:                                                      │")
        for k, v in rec.items():
            line = f"│    {k}: {str(v)}"
            preview_lines.append(f"{line:<64}│")

    preview_lines += [
        "├─────────────────────────────────────────────────────────────┤",
        "│  DECISIONS                                                   │",
        "│    [A] Approve  — save as-is                                │",
        "│    [R] Reject   — re-extract with feedback                  │",
        "│    [E] Edit     — paste corrected JSON records              │",
        "└─────────────────────────────────────────────────────────────┘",
    ]
    print("\n".join(preview_lines))

    # ── Pause and hand control back to the runner ─────────────────────────────
    human_response: dict = interrupt({
        "schema":  schema,
        "records": records[:5],
        "prompt":  "Enter decision: A (approve) / R (reject + feedback) / E (edit records)",
    })

    decision = human_response.get("decision", "approved").lower()

    if decision in ("a", "approve", "approved"):
        log.info("human_review: APPROVED")
        return Command(update={
            "review_decision": "approved",
            "messages": ["[human_review] decision=approved"],
        })

    if decision in ("r", "reject", "rejected"):
        feedback = human_response.get("feedback", "")
        log.info("human_review: REJECTED — feedback: %s", feedback)
        return Command(update={
            "review_decision": "rejected",
            "human_feedback":  feedback,
            "retries":         1,
            "messages": [f"[human_review] decision=rejected feedback='{feedback}'"],
        })

    if decision in ("e", "edit", "edited"):
        edited = human_response.get("records", records)
        log.info("human_review: EDITED — %d rows", len(edited))
        return Command(update={
            "review_decision": "edited",
            "edited_records":  edited,
            "messages": [f"[human_review] decision=edited rows={len(edited)}"],
        })

    # Fallback
    return Command(update={
        "review_decision": "approved",
        "messages": ["[human_review] fallback→approved"],
    })


# ── 5. apply_edits ────────────────────────────────────────────────────────────
def apply_edits(state: AgentState) -> dict:
    edited = state.get("edited_records") or state["records"]
    log.info("apply_edits: accepting %d edited rows", len(edited))
    return {
        "records":        edited,
        "edited_records": None,
        "messages": [f"[apply_edits] accepted {len(edited)} edited rows"],
    }


# ── 6. save_outputs ───────────────────────────────────────────────────────────
def save_outputs(state: AgentState) -> dict:
    schema  = state["schema"]
    records = state["records"]
    out_dir = Path(state["output_dir"])
    formats = state["formats"]
    ts      = datetime.now().strftime("%Y%m%d_%H%M%S")
    stem    = Path(state["source"]).stem if Path(state["source"]).exists() else "extracted"
    base    = f"{stem}_{ts}"
    out_dir.mkdir(parents=True, exist_ok=True)

    saved: dict[str, str] = {}

    # JSON
    if "json" in formats:
        p = out_dir / f"{base}.json"
        p.write_text(
            json.dumps({"schema": schema, "records": records},
                       indent=2, ensure_ascii=False, default=str),
            encoding="utf-8",
        )
        saved["json"] = str(p)
        log.info("Saved JSON  → %s", p)

    # CSV
    if "csv" in formats:
        p = out_dir / f"{base}.csv"
        with open(p, "w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=schema, extrasaction="ignore")
            w.writeheader()
            for rec in records:
                w.writerow({k: rec.get(k, "") for k in schema})
        saved["csv"] = str(p)
        log.info("Saved CSV   → %s", p)

    # Excel
    if "excel" in formats and EXCEL_OK:
        p  = out_dir / f"{base}.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Extracted Data"

        thin  = Side(style="thin", color="CCCCCC")
        bdr   = Border(left=thin, right=thin, top=thin, bottom=thin)
        hfill = PatternFill("solid", fgColor="1A56DB")
        hfont = Font(bold=True, color="FFFFFF", size=11, name="Calibri")
        halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
        altfill = PatternFill("solid", fgColor="EFF6FF")
        bfont   = Font(name="Calibri", size=10)
        balign  = Alignment(vertical="center")

        for ci, col in enumerate(schema, 1):
            c = ws.cell(row=1, column=ci, value=col.replace("_", " ").title())
            c.fill, c.font, c.alignment, c.border = hfill, hfont, halign, bdr

        for ri, rec in enumerate(records, 2):
            for ci, col in enumerate(schema, 1):
                c = ws.cell(row=ri, column=ci, value=rec.get(col))
                c.font, c.alignment, c.border = bfont, balign, bdr
                if ri % 2 == 0:
                    c.fill = altfill

        for col_cells in ws.columns:
            w_val = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[col_cells[0].column_letter].width = min(w_val + 4, 45)

        ws.freeze_panes    = "A2"
        ws.auto_filter.ref = ws.dimensions
        wb.save(p)
        saved["excel"] = str(p)
        log.info("Saved Excel → %s", p)

    return {
        "output_files": saved,
        "messages": [f"[save_outputs] wrote {list(saved.keys())}"],
    }


# ── 7. report ─────────────────────────────────────────────────────────────────
def report(state: AgentState) -> dict:
    summary = {
        "source":       state["source"][:100],
        "input_type":   state.get("input_type", "?"),
        "model":        GROQ_MODEL,
        "columns":      len(state.get("schema", [])),
        "records":      len(state.get("records", [])),
        "retries":      state.get("retries", 0),
        "decision":     state.get("review_decision"),
        "output_files": state.get("output_files", {}),
        "error":        state.get("extraction_error"),
        "timestamp":    datetime.now().isoformat(),
        "audit_log":    state.get("messages", []),
    }
    print("\n" + "═" * 60)
    print("  PIPELINE COMPLETE")
    print("═" * 60)
    print(f"  Source      : {summary['source']}")
    print(f"  Type        : {summary['input_type']}")
    print(f"  Model       : {summary['model']}")
    print(f"  Columns     : {summary['columns']}")
    print(f"  Records     : {summary['records']}")
    print(f"  Retries     : {summary['retries']}")
    print(f"  Decision    : {summary['decision']}")
    if summary["error"]:
        print(f"  ERROR       : {summary['error']}")
    if summary["output_files"]:
        print("\n  Saved files:")
        for fmt, path in summary["output_files"].items():
            print(f"    [{fmt.upper():6}]  {path}")
    print("\n  Audit log:")
    for entry in summary["audit_log"]:
        print(f"    {entry}")
    print("═" * 60 + "\n")
    return {"summary": summary}


# ── 8. report_error ───────────────────────────────────────────────────────────
def report_error(state: AgentState) -> dict:
    print(f"\n  [ERROR] Extraction failed: {state.get('extraction_error')}\n")
    return {"summary": {"error": state.get("extraction_error"), "records": 0}}


# ══════════════════════════════════════════════════════════════════════════════
#  CONDITIONAL ROUTING
# ══════════════════════════════════════════════════════════════════════════════

def route_by_type(state: AgentState) -> str:
    return f"load_{state['input_type']}"


def route_after_extraction(state: AgentState) -> str:
    return "report_error" if state.get("extraction_error") else "human_review"


def route_after_review(state: AgentState) -> str:
    decision = state.get("review_decision", "approved")
    if decision == "rejected":
        return "extract_data"
    if decision == "edited":
        return "apply_edits"
    return "save_outputs"


# ══════════════════════════════════════════════════════════════════════════════
#  GRAPH BUILDER
# ══════════════════════════════════════════════════════════════════════════════

def build_graph(checkpointer=None):
    g = StateGraph(AgentState)

    g.add_node("detect_type",  detect_type)
    g.add_node("load_image",   load_image)
    g.add_node("load_pdf",     load_pdf)
    g.add_node("load_audio",   load_audio)
    g.add_node("load_text",    load_text)
    g.add_node("extract_data", extract_data)
    g.add_node("human_review", human_review)
    g.add_node("apply_edits",  apply_edits)
    g.add_node("save_outputs", save_outputs)
    g.add_node("report",       report)
    g.add_node("report_error", report_error)

    g.set_entry_point("detect_type")

    g.add_conditional_edges(
        "detect_type", route_by_type,
        {"load_image": "load_image", "load_pdf": "load_pdf",
         "load_audio": "load_audio", "load_text": "load_text"},
    )

    for loader in ("load_image", "load_pdf", "load_audio", "load_text"):
        g.add_edge(loader, "extract_data")

    g.add_conditional_edges(
        "extract_data", route_after_extraction,
        {"human_review": "human_review", "report_error": "report_error"},
    )

    g.add_conditional_edges(
        "human_review", route_after_review,
        {"save_outputs": "save_outputs",
         "extract_data": "extract_data",
         "apply_edits":  "apply_edits"},
    )

    g.add_edge("apply_edits",  "save_outputs")
    g.add_edge("save_outputs", "report")
    g.add_edge("report",       END)
    g.add_edge("report_error", END)

    return g.compile(checkpointer=checkpointer)


# ══════════════════════════════════════════════════════════════════════════════
#  RUNNER
# ══════════════════════════════════════════════════════════════════════════════

MAX_RETRIES = 3


def run_pipeline(
    source:       str,
    api_key:      str | None = None,
    output_dir:   str        = "output",
    formats:      list[str]  = None,
    auto_approve: bool        = False,
    thread_id:    str | None = None,
) -> dict:
    """Execute the LangGraph pipeline with full HITL support."""
    if formats is None:
        formats = ["csv", "excel", "json"]

    key = api_key or os.environ.get("GROQ_API_KEY")
    if not key:
        raise ValueError(
            "Groq API key not found.\n\n"
            "  EASIEST FIX — create a .env file next to this script:\n"
            "    GROQ_API_KEY=gsk_your-key-here\n\n"
            "  OR set it in your terminal:\n"
            "    PowerShell : $env:GROQ_API_KEY='gsk_...'\n"
            "    CMD        : set GROQ_API_KEY=gsk_...\n"
            "    macOS/Linux: export GROQ_API_KEY='gsk_...'\n\n"
            "  OR pass it directly:\n"
            "    python multimodal_data_entry_assistant.py --api-key gsk_..."
        )

    checkpointer = MemorySaver()
    app          = build_graph(checkpointer=checkpointer)
    thread_id    = thread_id or str(uuid.uuid4())
    config       = {"configurable": {"thread_id": thread_id}}

    initial_state: AgentState = {
        "source":           source,
        "output_dir":       output_dir,
        "formats":          formats,
        "api_key":          key,
        "auto_approve":     auto_approve,
        "input_type":       "",
        "groq_messages":    [],
        "schema":           [],
        "records":          [],
        "extraction_error": None,
        "review_decision":  None,
        "human_feedback":   None,
        "edited_records":   None,
        "messages":         [],
        "retries":          0,
        "output_files":     {},
        "summary":          {},
    }

    result  = app.invoke(initial_state, config)
    retries = 0

    while True:
        snap = app.get_state(config)
        if not snap.next:
            break

        if snap.next[0] != "human_review":
            break

        retries += 1
        if retries > MAX_RETRIES:
            log.warning("Max retries (%d) reached — auto-approving.", MAX_RETRIES)
            result = app.invoke(Command(resume={"decision": "approved"}), config)
            break

        human_response = _collect_human_decision(snap.values)
        result = app.invoke(Command(resume=human_response), config)

    return result


def _collect_human_decision(state: dict) -> dict:
    """Terminal UI for collecting the human review decision."""
    while True:
        choice = input("\n  Decision [A/R/E]: ").strip().upper()

        if choice == "A":
            return {"decision": "approved"}

        if choice == "R":
            fb = input("  Feedback for re-extraction: ").strip()
            return {"decision": "rejected", "feedback": fb}

        if choice == "E":
            print("  Paste corrected JSON records array (end with a blank line):")
            lines = []
            while True:
                ln = input()
                if ln == "":
                    break
                lines.append(ln)
            try:
                edited = json.loads("\n".join(lines))
                if isinstance(edited, list):
                    return {"decision": "edited", "records": edited}
                print("  Must be a JSON array. Try again.")
            except json.JSONDecodeError as e:
                print(f"  Invalid JSON: {e}. Try again.")
        else:
            print("  Please enter A, R, or E.")


# ══════════════════════════════════════════════════════════════════════════════
#  INTERACTIVE REPL
# ══════════════════════════════════════════════════════════════════════════════

def interactive_mode(api_key, output_dir, formats, auto_approve):
    print("""
  ╔══════════════════════════════════════════════════════════╗
  ║  Multi-Modal Data Entry Assistant (LangGraph HITL)       ║
  ║  Groq  ·  llama-3.3-70b-versatile  ·  Whisper-large-v3  ║
  ╠══════════════════════════════════════════════════════════╣
  ║  Enter a file path  OR  paste text directly.             ║
  ║  Commands:  formats | output | auto | quit               ║
  ╚══════════════════════════════════════════════════════════╝
""")
    while True:
        try:
            raw = input("📥  Input > ").strip()
        except (EOFError, KeyboardInterrupt):
            print("\nBye!")
            break
        if not raw:
            continue
        cmd = raw.lower()
        if cmd in {"quit", "exit", "q"}:
            print("Bye!")
            break
        if cmd == "formats":
            print(f"  {formats}")
            new = input("  New (csv/excel/json comma-sep): ").strip()
            if new:
                formats = [f.strip() for f in new.split(",") if f.strip()]
            continue
        if cmd == "output":
            output_dir = input("  New output dir: ").strip() or output_dir
            continue
        if cmd == "auto":
            auto_approve = not auto_approve
            print(f"  auto_approve → {auto_approve}")
            continue
        try:
            run_pipeline(raw, api_key=api_key, output_dir=output_dir,
                         formats=formats, auto_approve=auto_approve)
        except Exception as exc:
            print(f"\n  ERROR: {exc}\n")
            log.debug("", exc_info=True)


# ══════════════════════════════════════════════════════════════════════════════
#  CLI
# ══════════════════════════════════════════════════════════════════════════════

def build_parser():
    p = argparse.ArgumentParser(
        prog="data_entry_assistant",
        description="Multi-Modal Data Entry Assistant — LangGraph HITL + Groq",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    p.add_argument("--input",        "-i", nargs="+", metavar="FILE")
    p.add_argument("--text",         "-t", metavar="TEXT")
    p.add_argument("--formats",      "-f", nargs="+",
                   default=["json", "csv", "excel"],
                   choices=["json", "csv", "excel"])
    p.add_argument("--output-dir",   "-o", default="output", metavar="DIR")
    p.add_argument("--api-key",      "-k", metavar="KEY",
                   help="Groq API key (overrides GROQ_API_KEY env var)")
    p.add_argument("--auto-approve", action="store_true",
                   help="Skip human review (fully autonomous mode).")
    p.add_argument("--verbose",      "-v", action="store_true")
    return p


def main():
    args = build_parser().parse_args()
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    if not args.input and not args.text:
        interactive_mode(args.api_key, args.output_dir,
                         args.formats, args.auto_approve)
        return

    sources = ([args.text] if args.text else []) + (args.input or [])
    for src in sources:
        try:
            run_pipeline(src, api_key=args.api_key, output_dir=args.output_dir,
                         formats=args.formats, auto_approve=args.auto_approve)
        except Exception as exc:
            print(f"\n  ERROR processing '{src}': {exc}\n")
            log.debug("", exc_info=True)


if __name__ == "__main__":
    main()
